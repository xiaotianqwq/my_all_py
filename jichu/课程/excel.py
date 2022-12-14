# -*- conding = utf-8 -*-
# @Time:2021/8/31 13:21
# @Author:宇
# @Flie:excel.py
# @Software:PyCharm
from openpyxl import Workbook,load_workbook

# wb = Workbook()
# sheet = wb.active  #获取当前active的sheet
#
# print(sheet.title)
# sheet.title = '123'
#
# #写数据，方式一
# sheet['D9'] = '123'
# sheet['C9'] = '456'
#
# #方式二
# sheet.append(['123','456','789'])
#
#
#
# wb.save('excel_text.xlsx')      #保存表

wb = load_workbook('excel_text.xlsx')   #打开excel
# print(wb.sheetnames)
sheet = wb['123']
# print(sheet['A5'].value)
# print(sheet['A5:A10'])
# for call in sheet['A5:A10']:
#     print(call[0].value)

# #全部循环
# for row in sheet:
#     #print(row)
#     for call in row:
#         print(call.value,end=' ')
#     print()

# #循环指定行
# for row in sheet.iter_rows(min_row=5,max_row=10):
#     #print(row)
#     for cell in row:
#         print(cell.value,end=' ')
#     print()

# #循环指定列
# for col in sheet.columns:
#     #print(col)
#     for cell in col:
#         print(cell.value,end=' ')
#     print()


# for col in sheet.iter_cols(min_col=5,max_col=10):
#     #print(col)
#     for cell in col:
#         print(cell.value,end=' ')
#     print()
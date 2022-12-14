# -*- conding = utf-8 -*-
# @Time:2022/3/15 16:36
# @Author:宇
# @Flie:img转表格.py
# @Software:PyCharm


from PIL import Image
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def get_size_RGB():
    # 图片路径
    img_file = '1.jpg'
    im = Image.open(img_file)
    # 图片格式转换
    out = im.convert('RGB')
    # 图片大小
    w, h = out.size
    # 像素RGB
    img_RGB = np.array(out)
    print(img_RGB[0][0])
    for i in img_RGB[0][0]:
        print(i)
    # print(img_RGB)
    return w, h, img_RGB


def rgb_hex(rgb):
    hex1 = ''
    for i in rgb:
        hex1 += str(hex(i))[2::]
    # print(hex1)
    return hex1


def make_woke(img_RGB,w,h):
    # 创建新的工作簿
    wb = Workbook()
    ws = wb.active
    # 循环单元格填充颜色
    for i in range(1,h+1):
        for j in range(1,w+1):
            gz = ws.cell(j,i)
            # print(j,i)
            hex1 = rgb_hex(img_RGB[i-1][j-1])
            # print(hex1)
            # ARGB?????????????
            fill = PatternFill("solid", fgColor='????????????????????????????????????????????????????????')
            gz.fill = fill

    wb_path = './1.xlsx'
    wb.save(wb_path)


def main():
    # 1.得到图片分辨率和RGB
    w, h, img_RGB = get_size_RGB()

    # 编辑单元格
    make_woke(img_RGB,w,h)


if __name__ == '__main__':
    main()
